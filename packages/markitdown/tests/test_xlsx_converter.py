#!/usr/bin/env python3 -m pytest
import io

import pytest
from openpyxl import Workbook

from markitdown import MarkItDown, XlsxConfig


def _convert(data: bytes, *, preserve_formatting: bool = False) -> str:
    md = MarkItDown(
        xlsx_config=XlsxConfig(preserve_formatting=preserve_formatting)
    )
    return md.convert(io.BytesIO(data), file_extension=".xlsx").text_content


def test_xlsx_default_does_not_preserve_formatting() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Cost"])
    ws.append(["Breakfast", 5])
    ws.cell(row=2, column=2).number_format = '"$"#,##0.00'
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue())
    assert "5" in result
    assert "$" not in result


def test_xlsx_preserve_currency() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Item", "Cost", "Total"])
    ws.append(["Breakfast", 5, 100])
    ws.append(["Laptops", 1199, 5995])
    for row in (2, 3):
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "$5.00" in result
    assert "$100.00" in result
    assert "$1,199.00" in result
    assert "$5,995.00" in result


def test_xlsx_preserve_percentage() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Rate"])
    ws.append(["A", 0.255])
    ws.cell(row=2, column=2).number_format = "0.0%"
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "25.5%" in result


def test_xlsx_preserve_thousands_separator() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Qty"])
    ws.append(["A", 1234567])
    ws.cell(row=2, column=2).number_format = "#,##0"
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "1,234,567" in result


def test_xlsx_preserve_other_currency_symbol() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Price"])
    ws.append(["A", 9.5])
    ws.cell(row=2, column=2).number_format = '"€"#,##0.00'
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "€9.50" in result


def test_xlsx_preserve_negative_currency() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Balance"])
    ws.append(["A", -1234.5])
    ws.cell(row=2, column=2).number_format = '"$"#,##0.00'
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "-$1,234.50" in result


def test_xlsx_preserve_dates() -> None:
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Date"])
    ws.append(["A", datetime(2024, 1, 15)])
    ws.cell(row=2, column=2).number_format = "yyyy-mm-dd"
    buf = io.BytesIO()
    wb.save(buf)

    result = _convert(buf.getvalue(), preserve_formatting=True)
    assert "2024-01-15" in result


def test_xlsx_config_exported() -> None:
    assert XlsxConfig(preserve_formatting=True).preserve_formatting is True
    assert XlsxConfig().preserve_formatting is False
