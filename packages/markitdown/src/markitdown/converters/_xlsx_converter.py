import re
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import BinaryIO, Any, Optional
from ._html_converter import HtmlConverter
from .._base_converter import DocumentConverter, DocumentConverterResult
from .._exceptions import MissingDependencyException, MISSING_DEPENDENCY_MESSAGE
from .._stream_info import StreamInfo

# Try loading optional (but in this case, required) dependencies
# Save reporting of any exceptions for later
_xlsx_dependency_exc_info = None
try:
    import pandas as pd
    import openpyxl  # noqa: F401
except ImportError:
    _xlsx_dependency_exc_info = sys.exc_info()

_xls_dependency_exc_info = None
try:
    import pandas as pd  # noqa: F811
    import xlrd  # noqa: F401
except ImportError:
    _xls_dependency_exc_info = sys.exc_info()

ACCEPTED_XLSX_MIME_TYPE_PREFIXES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
]
ACCEPTED_XLSX_FILE_EXTENSIONS = [".xlsx"]

ACCEPTED_XLS_MIME_TYPE_PREFIXES = [
    "application/vnd.ms-excel",
    "application/excel",
]
ACCEPTED_XLS_FILE_EXTENSIONS = [".xls"]


@dataclass
class XlsxConfig:
    """Configuration options for :class:`XlsxConverter`.

    Attributes:
        preserve_formatting: When ``True``, cell number formats (currency,
            percentage, thousands separators, dates, etc.) are preserved in the
            Markdown output. Defaults to ``False`` to keep the existing behavior
            of emitting raw cell values.
    """

    preserve_formatting: bool = False


class XlsxConverter(DocumentConverter):
    """
    Converts XLSX files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self, config: Optional[XlsxConfig] = None):
        super().__init__()
        self._html_converter = HtmlConverter()
        self._config = config or XlsxConfig()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLSX_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLSX_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:
        # Check the dependencies
        if _xlsx_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xlsx",
                    feature="xlsx",
                )
            ) from _xlsx_dependency_exc_info[
                1
            ].with_traceback(  # type: ignore[union-attr]
                _xlsx_dependency_exc_info[2]
            )

        if self._config.preserve_formatting:
            return self._convert_with_formatting(file_stream, **kwargs)

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="openpyxl")
        md_content = ""
        for s in sheets:
            md_content += f"## {s}\n"
            html_content = sheets[s].to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())

    def _convert_with_formatting(
        self,
        file_stream: BinaryIO,
        **kwargs: Any,
    ) -> DocumentConverterResult:
        """Convert an XLSX file to Markdown while preserving cell number formats."""
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        md_content = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            if sheet.max_row is None or sheet.max_column is None:
                continue

            md_content += f"## {sheet_name}\n"
            data = [
                [self._format_cell_value(cell) for cell in row]
                for row in sheet.iter_rows()
            ]
            if not data:
                continue

            df = pd.DataFrame(data)
            # Use the first row as the header, matching pandas.read_excel defaults
            header = [
                str(c) if c is not None and c != "" else f"Unnamed: {i}"
                for i, c in enumerate(df.iloc[0])
            ]
            df = df.iloc[1:]
            df.columns = header
            html_content = df.to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())

    def _format_cell_value(self, cell) -> str:
        """Render a cell value using its Excel number format."""
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, datetime):
            return self._format_datetime(value, cell.number_format)
        if not isinstance(value, (int, float)):
            return str(value)

        number_format = cell.number_format
        if not number_format or number_format == "General":
            return str(value)

        if "%" in number_format:
            return self._format_percentage(value, number_format)

        currency_symbol = self._extract_currency_symbol(number_format)
        if currency_symbol is not None:
            return self._format_currency(value, number_format, currency_symbol)

        if "#,##" in number_format or re.search(r"0+\.0+", number_format):
            return self._format_number(value, number_format)

        return str(value)

    def _extract_currency_symbol(self, number_format: str) -> Optional[str]:
        """Return the currency symbol embedded in an Excel number format, if any."""
        # Quoted literal, e.g. "$"#,##0.00 or #,##0.00"€"
        for match in re.finditer(r'"([^"]*)"', number_format):
            symbol = match.group(1).strip()
            if symbol and not symbol.replace(".", "").isdigit():
                return symbol
        # Currency code, e.g. [$USD] or [$¥-411]
        for match in re.finditer(r"\[\$([^\]]*)\]", number_format):
            code = re.sub(r"-\d+$", "", match.group(1)).strip()
            if code:
                return code
        return None

    def _format_number(self, value, number_format: str) -> str:
        decimal_places = self._decimal_places(number_format)
        if "#,##" in number_format:
            return f"{value:,.{decimal_places}f}"
        return f"{value:.{decimal_places}f}"

    def _format_percentage(self, value, number_format: str) -> str:
        decimal_places = self._decimal_places(number_format)
        return f"{value * 100:.{decimal_places}f}%"

    def _format_currency(
        self, value, number_format: str, symbol: str
    ) -> str:
        decimal_places = self._decimal_places(number_format)
        if "#,##" in number_format:
            formatted = f"{value:,.{decimal_places}f}"
        else:
            formatted = f"{value:.{decimal_places}f}"

        # Place the symbol before or after the number to match the format
        num_pos = len(number_format)
        for ch in "0#":
            idx = number_format.find(ch)
            if idx != -1:
                num_pos = min(num_pos, idx)
        symbol_pos = number_format.find(symbol)
        prefix = symbol_pos != -1 and symbol_pos < num_pos

        if value < 0:
            abs_formatted = formatted[1:]
            if "(" in number_format:
                return f"({symbol}{abs_formatted})"
            return (
                f"-{symbol}{abs_formatted}"
                if prefix
                else f"-{abs_formatted}{symbol}"
            )
        return f"{symbol}{formatted}" if prefix else f"{formatted}{symbol}"

    def _decimal_places(self, number_format: str) -> int:
        """Count the digit placeholders after the decimal point in a format."""
        if "." not in number_format:
            return 0
        decimal_places = 0
        for ch in number_format.split(".", 1)[1]:
            if ch == "0":
                decimal_places += 1
            elif ch in "#?":
                continue
            else:
                break
        return decimal_places

    def _format_datetime(self, value: datetime, number_format: str) -> str:
        fmt = number_format.lower()
        if any(tok in fmt for tok in ("hh", "h:", "am/pm")):
            return value.strftime("%Y-%m-%d %H:%M")
        return value.strftime("%Y-%m-%d")


class XlsConverter(DocumentConverter):
    """
    Converts XLS files to Markdown, with each sheet presented as a separate Markdown table.
    """

    def __init__(self):
        super().__init__()
        self._html_converter = HtmlConverter()

    def accepts(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> bool:
        mimetype = (stream_info.mimetype or "").lower()
        extension = (stream_info.extension or "").lower()

        if extension in ACCEPTED_XLS_FILE_EXTENSIONS:
            return True

        for prefix in ACCEPTED_XLS_MIME_TYPE_PREFIXES:
            if mimetype.startswith(prefix):
                return True

        return False

    def convert(
        self,
        file_stream: BinaryIO,
        stream_info: StreamInfo,
        **kwargs: Any,  # Options to pass to the converter
    ) -> DocumentConverterResult:
        # Load the dependencies
        if _xls_dependency_exc_info is not None:
            raise MissingDependencyException(
                MISSING_DEPENDENCY_MESSAGE.format(
                    converter=type(self).__name__,
                    extension=".xls",
                    feature="xls",
                )
            ) from _xls_dependency_exc_info[
                1
            ].with_traceback(  # type: ignore[union-attr]
                _xls_dependency_exc_info[2]
            )

        sheets = pd.read_excel(file_stream, sheet_name=None, engine="xlrd")
        md_content = ""
        for s in sheets:
            md_content += f"## {s}\n"
            html_content = sheets[s].to_html(index=False)
            md_content += (
                self._html_converter.convert_string(
                    html_content, **kwargs
                ).markdown.strip()
                + "\n\n"
            )

        return DocumentConverterResult(markdown=md_content.strip())
